from PIL import Image, ImageDraw, ImageFont
from datetime import date
import openpyxl
import os

from pandas import read_excel
path = "participation.xlsx"
inputWorkbook = openpyxl.load_workbook(path)

inputWorksheet = inputWorkbook.active
user = []
objects = {}
# Iterate over the rows
objects = {}
for i, row in enumerate(inputWorksheet.iter_rows(min_row=1, max_row=inputWorksheet.max_row, values_only=True)):
    objects['name'] = row[0]  # Assuming 'name' is in the first column (index 0)
    objects['team'] = row[1]
    objects['email'] = row[2]
    objects['department'] = row[3]
    objects['date'] = row[4]
    print(objects)



    user.append(objects)
    objects = {}

date1 = date.today().strftime('%Y-%m-%d')
for person in user:
    image = Image.open('participation.jpg')
    name = person['name']
    team = person['team']
    department = person['department']

   # font = ImageFont.truetype(
    #    f'{os.getcwd()}/Work.ttf', size=40)
    font = ImageFont.truetype("C:\\Windows\\FONTS\\Times.ttf", 20)





    draw = ImageDraw.Draw(image)
    draw.text(xy=(150, 262), text=name, fill=(0, 0, 0), font=font)
    draw.text(xy=(285, 295), text=team, fill=(0, 0, 0), font=font)
    draw.text(xy=(525, 262), text=department, fill=(0, 0, 0), font=font)
    draw.text(xy=(650, 295), text=date1, fill=(0, 0, 0), font=font)
    image.save(f'E:\Certification_Gen\\Certificates\\{name}_{team}.jpg')
    print(f'Generated for {name}_{team}')