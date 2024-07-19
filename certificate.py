import openpyxl
# import os
import smtplib
import ssl
import getpass
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

sender = "shivrajtestuse@gmail.com"
password = "lqpo sshu pfyy wixt"
subject = "SVSMD''s KKI Polytechnic 2022 Participation Certificate"

path = "participation.xlsx"
inputWorkbook = openpyxl.load_workbook(path)

inputWorksheet = inputWorkbook.active
user = []
objects = {}
# Iterate over the rows
objects = {}
for i, row in enumerate(inputWorksheet.iter_rows(min_row=1, max_row=inputWorksheet.max_row, values_only=True)):
    objects['name'] = row[0]  # Assuming 'name' is in the first column (index 0)
    objects['team'] = row[1]
    objects['department'] = row[2]
    objects['email'] = row[3]
#    objects['date'] = row[4]

    user.append(objects)
    objects = {}

context = ssl.create_default_context()

server = smtplib.SMTP("smtp.gmail.com", 587)
server.ehlo()
server.starttls(context=context)
print("Connected with Gmail Server")
server.ehlo()
server.login("shivrajtestuse@gmail.com","lqpo sshu pfyy wixt")
#server.login(sender, password)

for person in user:
    try:
        msg = MIMEMultipart()
        msg['From'] = sender
        msg['Subject'] = subject

        team = person['team']
        name = person['name']
        department = person['department']
        email = person['email']
        #date = person['date']

        body = f'Hi {name}, \n\nThank you for participating in the Quiz Competition 2022.'
        msg.attach(MIMEText(body, 'plain'))

        file = f'C:\\Users\\shivr\\Certification_Gen\\Certificates\\{name}_{team}.jpg'
        attachment = open(file, 'rb')
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(attachment.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', 'attachement; filename= ' + f'{name}_{team}.jpg')
        msg.attach(part)
        text = msg.as_string()
        server.sendmail(sender, email, text)
        print(f'Sent mail to {email}')
    except Exception as e:
        # print(e.__str__(), f'Could not send an email to {email}')
        print(e.__str__(), f'Could not send an email to')
server.quit()



